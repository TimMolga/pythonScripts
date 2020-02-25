import openpyxl

#wb = openpyxl.Workbook()
wb = openpyxl.load_workbook("transactions.xlsx")
print(wb.sheetnames)

sheet = wb["Sheet1"]

cell = sheet["a1"]
# print(cell.row)
# print(cell.column)
# print(cell.coordinate)
# cell = sheet.cell(row=1, column=1)
# print(sheet.max_row)
# print(sheet.max_column)

# for row in range(1, sheet.max_row + 1):
#     for column in range(1, sheet.max_column + 1):
#         cell = sheet.cell(row, column)
#         print(cell.value)

column = sheet["a"]
cells = sheet["a:c"]
cell_coord = sheet["a1:c3"]
cellRange = sheet[1:3]
print(cells)

sheet.append([1, 2, 3])
wb.save("transactions2.xlsx")
